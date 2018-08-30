# -*- coding: utf-8 -*-
"""
Created on Thu June 12th 2018

@author: Benson.Chen benson.chen@ap.jll.com
"""

import pandas as pd
from openpyxl import load_workbook
import validation as vd
from supplyscrapy import qichacha
from confidence import getConfidence
import db
import warnings


# Source-Site-(City)-LoadRound
sourcename = 'CM-West-CD-1'
# YYYYMMDDHH
timestamp = '20180803'
# File path
path = r'C:\Users\Benson.Chen\JLL\TDIM-GZ - Documents\Capforce\CM - WEST'
rawfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_RAW.xlsx'
rawfilepath = path + rawfilename
backupfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_BACKUP2.xlsx'
backupfilepath = path + backupfilename
reviewfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_REVIEW2.xlsx'
reviewfilepath = path + reviewfilename
scrapyfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_SCRAPY.xlsx'
scrapyfilepath = path + scrapyfilename
# backupfilepath =r'C:\Users\Benson.Chen\Desktop\test_com.xlsx'

contact_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'Name', 'First_Name', 'Last_Name', 'First_Name_CN', 'Last_Name_CN', 'Email', 'Phone', 'Mobile', 'Fax', 'Title', 'Contact_Address', 'City', 'State', 'Postal_Code', 'Country', 'Preferred_Language', 'Invest_Sectors', 'Investor_Purpose', 'Source_Company_ID', 'Comment', 'Reject_Reason', 'Load', 'db_New', 'vc_Deduplicate', 'vn_Lastname_CN', 'vn_Name_Swap', 'vn_Name_Space', 'vn_Name_Check', 've_Email_Format', 've_Email_Suffix', 've_Email_Domain', 've_Email_Check']
contact_load_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'First_Name', 'Last_Name', 'First_Name_CN', 'Last_Name_CN', 'Email', 'Phone', 'Mobile', 'Fax', 'Title', 'Contact_Address', 'District', 'City', 'State', 'Postal_Code', 'Country', 'Preferred_Language', 'Invest_Sectors', 'Investor_Purpose', 'Source_Company_ID', 'Comment', 'Load']
company_colnames = ['Source_ID', 'Parent_Name', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'District', 'City', 'State', 'Postal_Code', 'Country', 'Company_Type', 'Phone', 'Fax', 'Email', 'Website', 'Industry', 'Revenue', 'Employee', 'Full_Address', 'Comment', 'ComName_temp', 'State_Abbr', 'Load', 'db_New', 'vc_Deduplicate', 'vc_Address']
company_load_colnames = ['Source_ID', 'Parent_Name', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'District', 'City', 'State', 'Postal_Code', 'Country', 'Company_Type', 'Phone', 'Fax', 'Email', 'Website', 'Industry', 'Revenue', 'Employee', 'Full_Address', 'Comment', 'ComName_temp', 'Load']
company_dup_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'City', 'State', 'Phone', 'Website', 'Email', 'ComName_temp', 'vc_Deduplicate', 'db_New', 'Load', 'vc_Master_ID']
logs_columns = ['Source_ID', 'Entity_Type', 'Field', 'Action_Type', 'Log_From', 'Log_To']


def run(phrase):
    # Deduplicate companies, find common companies and contacts
    if phrase == 'p1':
        print('Phrase 1: Deduplicate companies, find common companies and contacts.')

        company_raw_list = pd.read_excel(rawfilepath, sheet_name='Company', sort=False, dtype=str)
        contact_raw_list = pd.read_excel(rawfilepath, sheet_name='Contact', sort=False, dtype=str)
        # Initialization
        company_init_list = vd.init_list(company_raw_list, company_colnames, 'Company')
        contact_init_list = vd.init_list(contact_raw_list, contact_colnames, 'Contact', sourcename, timestamp)
        # Deduplication within source data
        company_common_list, contact_common_list = vd.validate_common(company_init_list, contact_init_list)
        # Map state abbreviation and enrich state
        company_common_list = vd.map_state(company_common_list)
        company_duplicate_list, company_duplicate_full, company_dedup_list, contact_common_list = vd.dedup_company(company_common_list, contact_common_list)
        # Deduplication against staging table
        company_db_return = db.get_all(company_load_colnames, 'Company')
        if company_db_return.empty:
            company_existing_list = company_db_return
        else:
            company_dedup_list, company_existing_list = vd.dedup_comany_db(company_dedup_list, company_db_return)
        # Keep companies not duplicates
        # company_dedup_list = company_dedup_list[company_dedup_list['Load'] == True]

        print('Check {}, {}, deduplicate companies need review. {} contains full list of duplicate companies.'.format(reviewfilepath, '1_Duplicate_Company', '1_Duplicate_Company_Full'))
        print('{} companies are duplicates in this load.'.format(len(company_duplicate_list)))
        print('Check {}, {}.'.format(reviewfilepath, '2_Existing_Company'))
        print('{} companies already exists in local staging table.'.format(len(company_existing_list)))

        backupwriter = pd.ExcelWriter(backupfilepath, engine='openpyxl')
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        company_duplicate_list.sort_values(by=['ComName_temp']).to_excel(reviewwriter, index=False, header=True, columns=company_dup_colnames, sheet_name='1_Duplicate_Company')
        company_duplicate_full.sort_values(by=['ComName_temp']).to_excel(reviewwriter, index=False, header=True, columns=company_dup_colnames, sheet_name='1_Duplicate_Company_Full')
        company_existing_list.sort_values(by=['ComName_temp']).to_excel(reviewwriter, index=False, header=True, columns=list(company_existing_list), sheet_name='2_Existing_Company')
        company_dedup_list.to_excel(backupwriter, index=False, header=True, columns=company_colnames, sheet_name='company_dedup_list')
        contact_common_list.to_excel(backupwriter, index=False, header=True, columns=contact_colnames, sheet_name='contact_common_list')
        backupwriter.save()
        backupwriter.close()
        reviewwriter.save()
        reviewwriter.close()

    # Run web scraper to enrich company details
    elif phrase == 'p2':
        print('Phrase 2: Run web scraper to enrich company details.')

        company_dedup_list = pd.read_excel(backupfilepath, sheet_name='company_dedup_list', sort=False)
        # company_dedup_list = company_dedup_list[company_dedup_list['db_New'] != False]
        company_dedup_list = company_dedup_list[(company_dedup_list['db_New'] != False) & pd.isnull(company_dedup_list['Billing_Address'])]
        company_scrapy_return = qichacha(company_dedup_list, scrapyfilepath, 'company_scrapy_return')
        company_scrapy_return.to_excel(scrapyfilepath, index=False, header=True, columns=list(company_scrapy_return), sheet_name='company_scrapy_return')

    # Enrich companies with web scraper returns,  validate contact
    elif phrase == 'p3':
        print('Phrase 3: Enrich companies with web scraper returns')

        company_scrapy_return = pd.read_excel(scrapyfilepath, sheet_name='company_scrapy_return', sort=False)
        company_scrapy_return = vd.init_list(company_scrapy_return, list(company_scrapy_return))
        company_scrapy_return['Confidence'] = company_scrapy_return.apply(getConfidence, axis=1)
        company_scrapy_return['境外公司'] = company_scrapy_return['境外公司'].replace({0: False, 1: True})
        company_scrapy_return.to_excel(scrapyfilepath, index=False, header=True, columns=list(company_scrapy_return), sheet_name='company_scrapy_return')
        company_dedup_list = pd.read_excel(backupfilepath, sheet_name='company_dedup_list', sort=False)
        company_scrapy_list, company_scrapy_verify = vd.enrich_company(company_dedup_list, company_scrapy_return, company_colnames)
        company_scrapy_verify = company_scrapy_verify[(company_scrapy_verify['vc_Deduplicate'] == True) & (company_scrapy_verify['db_New'] == True)]
        print('Check {}, {}, enrich companies without address.'.format(reviewfilepath, '3_No_Address_Company'))
        print('{} companies remain no address.'.format(len(company_scrapy_verify)))
        print('Phrase 4: Validate contact.')

        # Validate contact
        contact_common_list = pd.read_excel(backupfilepath, sheet_name='contact_common_list', sort=False)
        contact_db_return = db.get_all(contact_load_colnames, 'Contact')
        if contact_db_return.empty:
            contact_dedup_list = contact_common_list
        else:
            contact_dedup_list = vd.dedup_contact_db(contact_common_list, contact_db_return)

        contact_validate_list = vd.validate_contacts(contact_dedup_list, contact_colnames, company_scrapy_list)
        contact_review_list = contact_validate_list[contact_validate_list['Load'] == False]
        contact_validate_list = contact_validate_list[contact_validate_list['Load'] == True]

        backupwriter = pd.ExcelWriter(backupfilepath, engine='openpyxl')
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        backupbook = load_workbook(backupwriter.path)
        reviewbook = load_workbook(reviewwriter.path)
        backupwriter.book = backupbook
        reviewwriter.book = reviewbook
        company_scrapy_list.to_excel(backupwriter, index=False, header=True, columns=company_colnames, sheet_name='company_scrapy_list')
        company_scrapy_verify.sort_values(by=['ComName_temp']).to_excel(reviewwriter, index=False, header=True, columns=company_colnames, sheet_name='3_No_Address_Company')
        contact_validate_list.sort_values(by=['First_Name', 'Last_Name', 'First_Name_CN', 'Last_Name_CN']).to_excel(backupwriter, index=False, header=True, columns=contact_load_colnames, sheet_name='contact_validate_list')
        contact_review_list.to_excel(reviewwriter, index=False, header=True, columns=contact_colnames, sheet_name='4_Validate_Contact')
        backupwriter.save()
        backupwriter.close()
        reviewwriter.save()
        reviewwriter.close()

        print('Check {}, {}, contacts need to review.'.format(reviewfilepath, '4_Validate_Contact'))
        print('{} contacts needs review'.format(len(contact_review_list)))

    # Enrich companies, contacts with business return, load company and contact into staging table
    elif phrase == 'p4':
        print('Phrase 5: Enrich companies')
        company_duplicate_review = pd.read_excel(reviewfilepath, sheet_name='1_Duplicate_Company', sort=False)
        company_existing_review = pd.read_excel(reviewfilepath, sheet_name='2_Existing_Company', sort=False)
        company_address_review = pd.read_excel(reviewfilepath, sheet_name='3_No_Address_Company', sort=False)
        company_scrapy_list = pd.read_excel(backupfilepath, sheet_name='company_scrapy_list', sort=False)
        company_load_list = vd.enrich_business(company_scrapy_list, company_duplicate_review)
        company_load_list = vd.enrich_business(company_load_list, company_existing_review)
        # Enrich companies without address
        company_load_list = vd.enrich_no_address(company_load_list, company_address_review)
        company_load_list = vd.enrich_business(company_load_list, company_address_review)
        company_min_drop = company_address_review[company_address_review['Load'] == False]
        print('Check {}, {}, companies cannot meet minimum standard.'.format(reviewfilepath, '5_Company_Drop'))
        print('{} companies are dropped'.format(len(company_min_drop)))

        # Merge deduplicate companies and format relative contacts
        print('Phrase 6: Merge deduplicate companies and clean relative contacts. Enrich contacts with business return.')
        contact_validate_review = pd.read_excel(reviewfilepath, sheet_name='4_Validate_Contact', sort=False)
        contact_validate_list = pd.read_excel(backupfilepath, sheet_name='contact_validate_list', sort=False)

        contact_load_list = vd.enrich_business(contact_validate_list, contact_validate_review)
        temp, contact_load_list = vd.dedup_fix(company_load_list, contact_load_list, company_duplicate_review)
        temp, contact_load_list = vd.dedup_fix(company_load_list, contact_load_list, company_existing_review)
        contact_load_list = vd.enrich_contact(company_load_list, contact_load_list, company_load_colnames)
        contact_load_list = contact_load_list[contact_load_list['Load'] == True]
        contact_no_company = contact_load_list[contact_load_list['Load'] == False]
        contact_no_company['Reject_Reason'] = 'No company;  '
        contact_min_list = contact_validate_list[contact_validate_list['Load'] == False]
        contact_min_list = contact_min_list.append(contact_no_company)
        print('Check {}, {}, contacts  cannot meet minimum standard.'.format(reviewfilepath, '5_Contact_Drop'))
        print('{} contacts are dropped, because companies are dropped.'.format(len(contact_min_list)))

        # Load company and contact into staging table
        print('Phrase 7:  Load company and contact into staging table.')
        db.load_staging(company_load_list, company_load_colnames, 'Company', sourcename, timestamp)
        db.load_staging(contact_load_list, contact_load_colnames, 'Contact', sourcename, timestamp)
        print('{} companies load into staging table.'.format(len(company_load_list)))
        print('{} contacts load into staging table.'.format(len(contact_load_list)))

        print('Phrase 8: Cross-check and log merge, deletion, modification record.')
        # Loading logs
        company_raw_list = pd.read_excel(rawfilepath, sheet_name='Company', sort=False)
        contact_raw_list = pd.read_excel(rawfilepath, sheet_name='Contact', sort=False)
        contact_raw_list['Source_ID'] = list(range(1, (len(contact_raw_list) + 1)))
        contact_raw_list['Source_ID'] = contact_raw_list['Source_ID'].apply(lambda x: sourcename + '_' + timestamp + '_' + 'Contact' + '_' + str(x))
        company_logs = vd.staging_log(company_raw_list, company_load_list, 'Company', logs_columns)
        db.load_staging(company_logs, logs_columns, 'Logs', sourcename, timestamp)
        contact_logs = vd.staging_log(contact_raw_list, contact_load_list, 'Contact', logs_columns)
        db.load_staging(contact_logs, logs_columns, 'Logs', sourcename, timestamp)

        # Loading summary
        company_duplicate_list = pd.read_excel(reviewfilepath, sheet_name='1_Duplicate_Company_Full', sort=False)
        company_existing_list = pd.read_excel(reviewfilepath, sheet_name='2_Existing_Company', sort=False)
        company_standard_list = pd.read_excel(reviewfilepath, sheet_name='3_No_Address_Company', sort=False)
        company_summary = vd.staging_summary('Company', company_raw_list, company_duplicate_list, company_existing_list, company_standard_list, company_load_list)
        db.load_staging(company_summary, list(company_summary), 'Summary', sourcename, timestamp)
        contact_validate_list = pd.read_excel(reviewfilepath, sheet_name='4_Validate_Contact', sort=False)
        contact_duplicate_list = contact_validate_list[contact_validate_list['vc_Deduplicate'] == False]
        contact_existing_list = contact_validate_list[contact_validate_list['db_New'] == False]
        contact_standard_list = contact_validate_list[contact_validate_list['Load'] == False]
        contact_summary = vd.staging_summary('Contact', contact_raw_list, contact_duplicate_list, contact_existing_list, contact_standard_list, contact_load_list)
        db.load_staging(contact_summary, list(contact_summary), 'Summary', sourcename, timestamp)
        # db.load_staging(company_scrapy_return, list(company_scrapy_return), 'Scrapy', sourcename, timestamp)

        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        reviewbook = load_workbook(reviewwriter.path)
        reviewwriter.book = reviewbook
        company_min_drop.to_excel(reviewwriter, index=False, header=True, columns=company_load_colnames, sheet_name='5_Company_Drop')
        contact_min_list.to_excel(reviewwriter, index=False, header=True, columns=company_load_colnames, sheet_name='5_Contact_Drop')
        company_load_list.to_excel(reviewwriter, index=False, header=True, columns=company_load_colnames, sheet_name='6_Company_Load')
        contact_load_list.to_excel(reviewwriter, index=False, header=True, columns=contact_load_colnames, sheet_name='6_Contact_Load')
        reviewwriter.save()
        reviewwriter.close()
        print('---------- Done.---------- ')


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    run('p3')

# contact_output =validate_contacts(contact_input_list, contact_colnames)
# contact_output.to_excel(r'C:\Users\Benson.Chen\Desktop\test.xlsx', index=False, header=True, columns=contact_colnames, sheet_name='Contact')
