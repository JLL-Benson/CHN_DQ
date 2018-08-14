# -*- coding: utf-8 -*-
"""
Created on Thu June 12th 2018

@author: Benson.Chen benson.chen@ap.jll.com
"""

import pandas as pd
from openpyxl import load_workbook
import validation as vd
import sys
import getpass
import dq
from supplyscrapy import qichacha
from confidence import getConfidence
import db
# sys.path.append(r'C:\Users\Benson.Chen\PycharmProjects\dq\jobs\live_sf\bau')
# import bau_cf_accounts_lib as bau_accounts
# import bau_cf_contacts as bau_contacts

# Source-Site-LoadRound
sourcename = 'CM-West-CQ-1'
# YYYYMMDDHH
timestamp = '20180803'
# File path
path = r'C:\Users\Benson.Chen\JLL\TDIM-GZ - Documents\Capforce\CM - WEST'
rawfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_RAW.xlsx'
rawfilepath = path + rawfilename
backupfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_BACKUP.xlsx'
backupfilepath = path + backupfilename
reviewfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_REVIEW.xlsx'
reviewfilepath = path + reviewfilename
scrapyfilename = r'\CHN-DQ_' + sourcename + '_' + timestamp + '_SCRAPY.xlsx'
scrapyfilepath = path + scrapyfilename
# backupfilepath =r'C:\Users\Benson.Chen\Desktop\test_com.xlsx'

contact_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'Name', 'First_Name', 'Last_Name', 'First_Name_CN', 'Last_Name_CN', 'Email', 'Phone', 'Mobile', 'Fax', 'Title', 'Source_Company_ID', 'Load', 'Reject_Reason', 'First_Name2', 'Last_Name2', 'Email2', 'vc_Deduplicate', 'vn_Lastname_CN', 'vn_Name_Swap', 'vn_Name_Space', 'vn_Name_Check', 've_Email_Format', 've_Email_Suffix', 've_Email_Domain', 've_Email_Check', 'db_New']
contact_load_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'First_Name', 'Last_Name', 'First_Name_CN', 'Last_Name_CN', 'Email', 'Phone', 'Mobile', 'Fax', 'Title', 'Source_Company_ID', 'Load']
company_colnames = ['Source_ID', 'Parent_Name', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'Postal_Code', 'District', 'City', 'State', 'Country', 'Company_Type', 'Phone', 'Fax', 'Email', 'Website', 'Industry', 'Revenue', 'Employee', 'Full_Address', 'ComName_temp', 'Load', 'db_New', 'vc_Address']
company_load_colnames = ['Source_ID', 'Parent_Name', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'Postal_Code', 'District', 'City', 'State', 'Country', 'Company_Type', 'Phone', 'Fax', 'Email', 'Website', 'Industry', 'Revenue', 'Employee', 'Full_Address', 'ComName_temp', 'Load']
company_dup_colnames = ['Source_ID', 'Company_Name', 'Company_Name_CN', 'Billing_Address', 'City', 'State', 'Phone', 'Website', 'Email', 'vc_Deduplicate', 'ComName_temp', 'Load', 'vc_Master_ID']
logs_columns = ['Source_ID', 'Entity_Type', 'Field', 'Action_Type', 'Log_From', 'Log_To']


def run(phrase):
    # Deduplicate companies, find common companies and contacts
    if phrase == 'p1':
        print('Phrase 1: Deduplicate companies, find common companies and contacts.')
        company_raw_list = pd.read_excel(rawfilepath, sheet_name='Company', sort=False, dtype=str)
        contact_raw_list = pd.read_excel(rawfilepath, sheet_name='Contact', sort=False, dtype=str)
        company_init_list = vd.init_list(company_raw_list, company_colnames, 'Company')
        contact_init_list = vd.init_list(contact_raw_list, contact_colnames, 'Contact', sourcename, timestamp)
        company_common_list, contact_common_list = vd.validate_common(company_init_list, contact_init_list)
        company_duplicate_list, company_duplicate_full, company_common_list, contact_common_list = vd.dedup_company(company_common_list, contact_common_list)
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        backupwriter = pd.ExcelWriter(backupfilepath, engine='openpyxl')
        company_duplicate_list.to_excel(reviewwriter, index=False, header=True, columns=company_dup_colnames, sheet_name='1_Duplicate_Company')
        company_duplicate_full.to_excel(reviewwriter, index=False, header=True, columns=company_dup_colnames, sheet_name='1_Duplicate_Company_Full')
        company_common_list.to_excel(backupwriter, index=False, header=True, columns=company_colnames, sheet_name='company_common_list')
        contact_common_list.to_excel(backupwriter, index=False, header=True, columns=contact_colnames, sheet_name='contact_common_list')
        reviewwriter.save()
        reviewwriter.close()
        backupwriter.save()
        backupwriter.close()
        print('Check {}, {}, deduplicate company'.format(backupfilepath, '1_Duplicate_Company'))

    # Merge deduplicate companies and format relative contacts
    elif phrase == 'p2':
        print('Phrase 2: Merge deduplicate companies and clean relative contacts.')
        company_common_list = pd.read_excel(backupfilepath, sheet_name='company_common_list', sort=False)
        contact_common_list = pd.read_excel(backupfilepath, sheet_name='contact_common_list', sort=False)
        company_duplicate_list = pd.read_excel(reviewfilepath, sheet_name='1_Duplicate_Company', sort=False)
        company_dedup_list, contact_format_list = vd.dedup_fix(company_common_list, contact_common_list, company_duplicate_list)
        company_db_return = db.get_all(company_load_colnames, 'Company')
        if company_db_return.empty:
            company_existing_list = company_db_return
        else:
            company_dedup_list, company_existing_list = vd.dedup_comany_db(company_dedup_list, company_db_return)

        backupwriter = pd.ExcelWriter(backupfilepath, engine='openpyxl')
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        backupbook = load_workbook(backupwriter.path)
        reviewbook = load_workbook(reviewwriter.path)
        backupwriter.book = backupbook
        reviewwriter.book = reviewbook
        company_existing_list.to_excel(reviewwriter, index=False, header=True, columns=company_colnames, sheet_name='2_Existing_Company')
        company_dedup_list.to_excel(backupwriter, index=False, header=True, columns=company_colnames, sheet_name='company_dedup_list')
        contact_format_list.to_excel(backupwriter, index=False, header=True, columns=contact_colnames, sheet_name='contact_format_list')
        backupwriter.save()
        backupwriter.close()
        reviewwriter.save()
        reviewwriter.close()

        print('Run p3')

    # Run web scraper to enrich company details
    elif phrase == 'p3':
        print('Phrase 3: Run web scraper to enrich company details.')
        company_dedup_list = pd.read_excel(backupfilepath, sheet_name='company_dedup_list', sort=False)
        company_dedup_list = vd.map_state(company_dedup_list)
        company_scrapy_return = qichacha(company_dedup_list[company_dedup_list['dq_New'] != False], scrapyfilepath, 'company_scrapy_return')
        company_scrapy_return.to_excel(scrapyfilepath, index=False, header=True, columns=list(company_scrapy_return), sheet_name='company_scrapy_return')

    elif phrase == 'p6':
        print('next step')
        company_scrapy_return = pd.read_excel(scrapyfilepath, sheet_name='company_scrapy_return', sort=False)
        company_scrapy_return = vd.init_list(company_scrapy_return, list(company_scrapy_return))
        company_scrapy_return['Confidence'] = company_scrapy_return.apply(getConfidence, axis=1)
        company_scrapy_return['境外公司'] = company_scrapy_return['境外公司'].replace({0: False, 1: True})
        company_scrapy_return.to_excel(scrapyfilepath, index=False, header=True, columns=list(company_scrapy_return), sheet_name='company_scrapy_return')
        company_dedup_list = pd.read_excel(backupfilepath, sheet_name='company_dedup_list', sort=False)
        company_scrapy_list, company_scrapy_verify = vd.enrich_company(company_dedup_list, company_scrapy_return, company_colnames)

        backupwriter = pd.ExcelWriter(backupfilepath, engine='openpyxl')
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        backupbook = load_workbook(backupwriter.path)
        reviewbook = load_workbook(reviewwriter.path)
        backupwriter.book = backupbook
        reviewwriter.book = reviewbook
        company_scrapy_list.to_excel(backupwriter, index=False, header=True, columns=company_colnames, sheet_name='company_scrapy_list')
        company_scrapy_verify.to_excel(reviewwriter, index=False, header=True, columns=company_colnames, sheet_name='3_No_Address_Company')
        backupwriter.save()
        backupwriter.close()
        reviewwriter.save()
        reviewwriter.close()
    # Enrich company with business return, validate contact
    elif phrase == 'p7':
        company_business_return = pd.read_excel(reviewfilepath, sheet_name='3_No_Address_Company', sort=False)
        company_scrapy_list = pd.read_excel(backupfilepath, sheet_name='company_scrapy_list', sort=False)
        contact_format_list = pd.read_excel(backupfilepath, sheet_name='contact_format_list', sort=False)
        company_load_list = vd.enrich_business(company_scrapy_list, company_business_return)
        contact_db_return = db.get_all(contact_load_colnames, 'Contact')
        if contact_db_return.empty:
            contact_dedup_list = contact_format_list
        else:
            contact_dedup_list = vd.dedup_contact_db(contact_format_list, contact_db_return)

        contact_validate_list = vd.validate_contacts(contact_dedup_list, contact_colnames, company_load_list)
        company_load_list = company_load_list[company_load_list['Load'] == True]
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        reviewbook = load_workbook(reviewwriter.path)
        reviewwriter.book = reviewbook
        contact_validate_list.to_excel(reviewwriter, index=False, header=True, columns=contact_colnames, sheet_name='4_Validate_Contact')
        company_load_list.to_excel(reviewwriter, index=False, header=True, columns=company_load_colnames, sheet_name='5_Company_Load')
        reviewwriter.save()
        reviewwriter.close()
        db.load_staging(company_load_list, company_load_colnames, 'Company', sourcename, timestamp)
    elif phrase == 'p8':
        contact_format_list = pd.read_excel(backupfilepath, sheet_name='contact_validate_list', sort=False)
        contact_business_list = pd.read_excel(reviewfilepath, sheet_name='4_Validate_Contact', sort=False)
        droplist = list(contact_business_list.loc[contact_business_list['Load'] == False, 'Source_ID'])
        contact_load_list = contact_format_list[~contact_format_list['Source_ID'].isin(droplist)]
        reviewwriter = pd.ExcelWriter(reviewfilepath, engine='openpyxl')
        reviewbook = load_workbook(reviewwriter.path)
        reviewwriter.book = reviewbook
        contact_load_list.to_excel(reviewwriter, index=False, header=True, columns=contact_load_colnames, sheet_name='5_Contact_Load')
        reviewwriter.save()
        reviewwriter.close()
        db.load_staging(contact_load_list, contact_load_colnames, 'Contact', sourcename, timestamp)



if __name__ == '__main__':
    run('p8')

# contact_output =validate_contacts(contact_input_list, contact_colnames)
# contact_output.to_excel(r'C:\Users\Benson.Chen\Desktop\test.xlsx', index=False, header=True, columns=contact_colnames, sheet_name='Contact')
